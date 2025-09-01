"""
ABC사 제품 재료비 트렌드 분석 엑셀 파일 데이터 생성기
1줄로 구성된 제목과 500개 모델의 재료비 데이터를 생성합니다.
매월 말일에 추출하여 재료비 트렌드를 분석하는 테이블입니다.
"""

import random
from datetime import datetime, date
import calendar
import os

class CostTrendsGenerator:
    def __init__(self):
        # 제품 시리즈
        self.series = [
            "A1", "A2", "B1", "B2", "C1", "C2", "D1", "D2", "E1", "E2"
        ]
        
        # 지역 옵션
        self.regions = ["북미", "국내", "독일", "영국"]
        
        # 인치 사이즈 (3~18인치, 15종류)
        self.inch_sizes = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18][:15]
        
        # 부품별 재료비 범위 (원 단위)
        self.component_ranges = {
            "Cell": (800.0, 1200.0),
            "BLU": (250.0, 400.0), 
            "광원": (150.0, 300.0),
            "광학": (80.0, 150.0),
            "BLU회로": (30.0, 80.0),
            "LCM기구": (40.0, 90.0),
            "방열": (0.3, 1.2),
            "회로": (400.0, 800.0),
            "Main": (150.0, 300.0),
            "OC/WOC box": (10.0, 25.0),
            "SMPS": (150.0, 250.0),
            "SPK": (180.0, 220.0),
            "회로기타": (200.0, 400.0),
            "SET기구": (80.0, 150.0)
        }
        
        # 월별 추출 날짜 생성 (2025년 2월~7월, 매월 말일)
        self.extract_dates = self.generate_monthly_dates()
    
    def generate_monthly_dates(self):
        """매월 말일 추출 날짜 생성 (2025년 2월~7월)"""
        dates = []
        current_year = 2025
        for month in range(2, 8):  # 2월부터 7월까지
            last_day = calendar.monthrange(current_year, month)[1]
            extract_date = date(current_year, month, last_day)
            dates.append(extract_date.strftime("%Y년 %m월 %d일"))
        return dates
    
    def generate_model_name(self, index):
        """모델명 생성 (model-A, model-B 형식)"""
        series = random.choice(self.series)
        model_num = f"{index:03d}"
        return f"model-{series}-{model_num}"
    
    def generate_component_costs(self, inch_size, series):
        """부품별 재료비 생성"""
        costs = {}
        
        # 인치 크기와 시리즈에 따른 가격 보정 계수
        size_multiplier = 0.8 + (inch_size - 3) * 0.02  # 3인치 기준 0.8, 18인치 기준 1.1
        series_multiplier = 1.0 + (ord(series[0]) - ord('A')) * 0.1  # A=1.0, B=1.1, C=1.2 등
        
        for component, (min_cost, max_cost) in self.component_ranges.items():
            base_cost = random.uniform(min_cost, max_cost)
            adjusted_cost = base_cost * size_multiplier * series_multiplier
            costs[component] = round(adjusted_cost, 1)
        
        return costs
    
    def read_existing_file(self, filename="03_cost_trends.xlsx"):
        """기존 엑셀 파일 읽기"""
        import openpyxl
        
        file_path = filename  # 현재 폴더에서 찾기
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except FileNotFoundError:
            return None

    def copy_header_structure(self, source_worksheet, target_worksheet):
        """원본 워크시트의 헤더 구조를 대상 워크시트로 복사"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 1행의 헤더 구조 복사
        max_col = 6 + len(self.component_ranges) + 1  # Model, 추출날짜, 지역, 인치, 전체재료비, 핵심 + 부품들 + 시리즈
        for col_num in range(1, max_col + 1):
            source_cell = source_worksheet.cell(row=1, column=col_num)
            target_cell = target_worksheet.cell(row=1, column=col_num)
            
            # 셀 값 복사
            target_cell.value = source_cell.value
            
            # 스타일 복사 (폰트, 정렬 등)
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical
                )

    def generate_data_row(self, model_index, extract_date):
        """데이터 행 생성 - 실제 엑셀 구조에 맞춘 데이터"""
        model_name = self.generate_model_name(model_index)
        region = random.choice(self.regions)
        inch_size = random.choice(self.inch_sizes)
        series = random.choice(self.series)
        
        # 부품별 재료비 생성
        component_costs = self.generate_component_costs(inch_size, series)
        
        # 핵심 재료비 (주요 부품들의 합계)
        core_components = ["Cell", "BLU", "광원", "광학", "Main"]
        core_cost = sum(component_costs.get(comp, 0) for comp in core_components)
        
        # 전체 재료비 (모든 부품의 합계)
        total_cost = sum(component_costs.values())
        
        # 기본 행 데이터 구성
        row = [
            model_name,      # Model
            extract_date,    # 추출날짜
            region,          # 지역
            inch_size,       # 인치
            round(total_cost, 1),  # 전체재료비(Total)
            round(core_cost, 1)    # 핵심
        ]
        
        # 각 부품별 재료비 추가
        for component in self.component_ranges.keys():
            row.append(component_costs[component])
        
        # 시리즈 추가
        row.append(series)
        
        return row

    def create_excel_file(self, filename="03_cost_trends.xlsx"):
        """기존 엑셀 파일 구조를 유지하면서 새 파일 생성"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # 원본 파일 경로와 새 파일 경로 설정
        original_path = filename  # 현재 폴더의 원본 파일
        new_filename = filename.replace(".xlsx", "_generated.xlsx")
        output_path = new_filename  # 현재 폴더에 새 파일 생성
        
        # 기존 파일 로드 시도
        existing_workbook = self.read_existing_file(filename)
        
        if existing_workbook:
            # 기존 파일이 있는 경우 - 헤더 구조 완전 복사
            original_worksheet = existing_workbook.active
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = original_worksheet.title
            
            # 헤더 구조 복사
            self.copy_header_structure(original_worksheet, worksheet)
            
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
            
        else:
            # 원본 파일이 없는 경우 - 기본 구조 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "재료비트렌드"
            
            # 1줄 헤더 생성
            self.create_default_header(worksheet)
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
        
        # 데이터 생성 및 입력 (50개 모델 × 6개월 = 300개 데이터)
        row_index = 0
        for extract_date in self.extract_dates:  # 6개월 반복
            for model_index in range(1, 51):  # 각 월별로 50개 모델
                row_data = self.generate_data_row(model_index, extract_date)
                row_num = start_row + row_index
                
                for col, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
                
                row_index += 1
        
        # 컬럼 너비 자동 조정 (병합된 셀 오류 방지)
        max_col = 6 + len(self.component_ranges) + 1
        for col_num in range(1, max_col + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # 성능을 위해 100행까지만 체크
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value and not isinstance(cell, MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 8, 최대 25)
            adjusted_width = min(max(max_length + 2, 8), 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_generated_{alt_timestamp}.xlsx")
            output_path = alt_filename
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_default_header(self, worksheet):
        """기본 1줄 헤더 생성 - 실제 이미지 구조 반영"""
        headers = [
            "Model", "추출날짜", "지역", "인치", "전체재료비(Total)", "핵심"
        ]
        
        # 부품별 재료비 컬럼 추가
        for component in self.component_ranges.keys():
            headers.append(component)
        
        # 시리즈 컬럼 추가
        headers.append("시리즈")
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

    def create_refined_excel_file(self, filename="03_cost_trends.xlsx"):
        """refined 버전의 엑셀 파일 생성 (1줄 헤더)"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        new_filename = filename.replace(".xlsx", "_refined.xlsx")
        output_path = os.path.join(refined_folder, new_filename)
        
        # 새 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "재료비트렌드_정제"
        
        # 1줄 헤더 생성
        headers = [
            "모델명", "추출날짜", "지역", "인치사이즈", "전체재료비", "핵심재료비"
        ]
        
        # 부품별 재료비 컬럼 추가
        for component in self.component_ranges.keys():
            headers.append(f"{component}_재료비")
        
        headers.append("제품시리즈")
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 데이터 생성 및 입력 (50개 모델 × 6개월 = 300개 데이터)
        row_index = 0
        for extract_date in self.extract_dates:  # 6개월 반복
            for model_index in range(1, 51):  # 각 월별로 50개 모델
                row_data = self.generate_data_row(model_index, extract_date)
                row_num = row_index + 2  # 헤더 다음 행부터
                
                for col, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
                
                row_index += 1
        
        # 컬럼 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # 성능을 위해 100행까지만 체크
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 10, 최대 20)
            adjusted_width = min(max(max_length + 2, 10), 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_refined_{alt_timestamp}.xlsx")
            output_path = os.path.join(refined_folder, alt_filename)
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 정제된 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_description_md_file(self):
        """설명 마크다운 파일 생성"""
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        filename = "03_cost_trends_description.md"
        output_path = os.path.join(refined_folder, filename)
        
        md_content = """# ABC사 제품 재료비 트렌드 분석 데이터베이스 명세서

## 1. SQLite 테이블 생성 SQL

```sql
CREATE TABLE product_cost_trends (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    model_name VARCHAR(50) NOT NULL,
    extract_date VARCHAR(20) NOT NULL,
    region VARCHAR(10) NOT NULL,
    inch_size INTEGER NOT NULL,
    total_cost DECIMAL(10,2) NOT NULL,
    core_cost DECIMAL(10,2) NOT NULL,
    cell_cost DECIMAL(8,2),
    blu_cost DECIMAL(8,2),
    light_source_cost DECIMAL(8,2),
    optical_cost DECIMAL(8,2),
    blu_circuit_cost DECIMAL(8,2),
    lcm_mechanism_cost DECIMAL(8,2),
    heat_dissipation_cost DECIMAL(8,2),
    circuit_cost DECIMAL(8,2),
    main_cost DECIMAL(8,2),
    oc_woc_box_cost DECIMAL(8,2),
    smps_cost DECIMAL(8,2),
    spk_cost DECIMAL(8,2),
    circuit_etc_cost DECIMAL(8,2),
    set_mechanism_cost DECIMAL(8,2),
    series VARCHAR(10) NOT NULL,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- 월별 트렌드 분석을 위한 뷰
CREATE VIEW monthly_cost_trends AS
SELECT 
    substr(extract_date, 1, 7) as year_month,
    region,
    series,
    AVG(total_cost) as avg_total_cost,
    AVG(core_cost) as avg_core_cost,
    COUNT(*) as model_count
FROM product_cost_trends 
GROUP BY substr(extract_date, 1, 7), region, series;

-- 인덱스 생성
CREATE INDEX idx_model_name ON product_cost_trends(model_name);
CREATE INDEX idx_extract_date ON product_cost_trends(extract_date);
CREATE INDEX idx_region ON product_cost_trends(region);
CREATE INDEX idx_series ON product_cost_trends(series);
CREATE INDEX idx_inch_size ON product_cost_trends(inch_size);
CREATE INDEX idx_total_cost ON product_cost_trends(total_cost);
```

## 2. 테이블 및 컬럼 상세 설명

### 테이블 개요
- **테이블명**: `product_cost_trends`
- **목적**: ABC사 제품의 월별 재료비 변동 추이를 관리하고 분석
- **데이터 구조**: 50개 제품 모델 × 6개월(2025년 2월~7월) = 총 300건 데이터
- **추출 방식**: 매월 말일 기준, 해당 월의 모든 모델이 동일한 추출날짜 보유

### 컬럼 상세 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `model_name` | VARCHAR(50) | 제품 모델명 | 'model-A1-001', 'model-B2-150' |
| `extract_date` | VARCHAR(20) | 데이터 추출 날짜 (매월 말일) | '2025년 07월 31일' |
| `region` | VARCHAR(10) | 판매/생산 지역 | '북미', '국내', '독일', '영국' |
| `inch_size` | INTEGER | 제품 인치 크기 | 3, 7, 12, 18 |
| `total_cost` | DECIMAL(10,2) | 전체 재료비 합계 | 2111.10, 3123.50 |
| `core_cost` | DECIMAL(10,2) | 핵심 부품 재료비 합계 | 2000.1, 1000.2 |
| `cell_cost` | DECIMAL(8,2) | Cell 부품 재료비 | 1005.7 |
| `blu_cost` | DECIMAL(8,2) | BLU 부품 재료비 | 300.1 |
| `light_source_cost` | DECIMAL(8,2) | 광원 부품 재료비 | 200.8 |
| `optical_cost` | DECIMAL(8,2) | 광학 부품 재료비 | 111.1 |
| `blu_circuit_cost` | DECIMAL(8,2) | BLU회로 부품 재료비 | 50.1 |
| `lcm_mechanism_cost` | DECIMAL(8,2) | LCM기구 부품 재료비 | 55.1 |
| `heat_dissipation_cost` | DECIMAL(8,2) | 방열 부품 재료비 | 0.5 |
| `circuit_cost` | DECIMAL(8,2) | 회로 부품 재료비 | 599.1 |
| `main_cost` | DECIMAL(8,2) | Main 부품 재료비 | 200.1 |
| `oc_woc_box_cost` | DECIMAL(8,2) | OC/WOC box 부품 재료비 | 15.1 |
| `smps_cost` | DECIMAL(8,2) | SMPS 부품 재료비 | 200.3 |
| `spk_cost` | DECIMAL(8,2) | SPK 부품 재료비 | 200.1 |
| `circuit_etc_cost` | DECIMAL(8,2) | 회로기타 부품 재료비 | 300.1 |
| `set_mechanism_cost` | DECIMAL(8,2) | SET기구 부품 재료비 | 102.5 |
| `series` | VARCHAR(10) | 제품 시리즈 | 'A1', 'B2', 'C1' |

## 3. 예상 질문과 SQL 쿼리

### 자주 묻는 질문들과 해당 SQL 쿼리

| 번호 | 질문 | SQL 쿼리 | 예상 결과 |
|------|------|----------|----------|
| 1 | model-A1-001 모델 재료비 얼마야? | `SELECT total_cost FROM product_cost_trends WHERE model_name = 'model-A1-001' ORDER BY extract_date DESC LIMIT 1;` | 2111.10 |
| 2 | model-B2-150 모델의 Cell 재료비는 얼마야? | `SELECT cell_cost FROM product_cost_trends WHERE model_name = 'model-B2-150' ORDER BY extract_date DESC LIMIT 1;` | 1005.7 |
| 3 | A1 시리즈의 연차별 재료비 알려줘 | `SELECT substr(extract_date, 1, 4) as year, AVG(total_cost) as avg_cost FROM product_cost_trends WHERE series = 'A1' GROUP BY substr(extract_date, 1, 4);` | 2025: 2500.30 등 |
| 4 | model-C1-200 모델의 5월 대비 7월 재료비 차분 알려줘 | `SELECT (jul.total_cost - may.total_cost) as cost_diff FROM product_cost_trends may JOIN product_cost_trends jul ON may.model_name = jul.model_name WHERE may.model_name = 'model-C1-200' AND may.extract_date LIKE '%05월%' AND jul.extract_date LIKE '%07월%';` | +150.20 |
| 5 | 북미 지역 평균 재료비는? | `SELECT AVG(total_cost) FROM product_cost_trends WHERE region = '북미';` | 2845.60 |
| 6 | Cell 부품 재료비가 가장 높은 모델은? | `SELECT model_name, cell_cost FROM product_cost_trends ORDER BY cell_cost DESC LIMIT 5;` | model-E2-450: 1200.0 등 |
| 7 | 6월 대비 7월 재료비 상승률 상위 10개 모델은? | `SELECT model_name, ((jul.total_cost - jun.total_cost) / jun.total_cost * 100) as increase_rate FROM product_cost_trends jun JOIN product_cost_trends jul ON jun.model_name = jul.model_name WHERE jun.extract_date LIKE '%06월%' AND jul.extract_date LIKE '%07월%' ORDER BY increase_rate DESC LIMIT 10;` | model-D1-300: +15.2% 등 |
| 8 | 15인치 이상 모델의 평균 재료비는? | `SELECT AVG(total_cost) FROM product_cost_trends WHERE inch_size >= 15;` | 4520.80 |
| 9 | 시리즈별 월별 재료비 트렌드는? | `SELECT series, substr(extract_date, 6, 2) as month, AVG(total_cost) FROM product_cost_trends GROUP BY series, substr(extract_date, 6, 2) ORDER BY series, month;` | A1 02월: 2300, A1 03월: 2350 등 |
| 10 | 재료비 이상치(평균대비 ±20% 초과) 모델은? | `WITH avg_cost AS (SELECT AVG(total_cost) as avg_val FROM product_cost_trends) SELECT model_name, total_cost FROM product_cost_trends, avg_cost WHERE total_cost > avg_val * 1.2 OR total_cost < avg_val * 0.8;` | model-E2-500: 5200.0 등 |

## 4. 데이터 활용 가이드

### 월별 트렌드 분석
```sql
-- 모델별 월별 재료비 변화 추이
SELECT 
    model_name,
    extract_date,
    total_cost,
    LAG(total_cost) OVER (PARTITION BY model_name ORDER BY extract_date) as prev_cost,
    total_cost - LAG(total_cost) OVER (PARTITION BY model_name ORDER BY extract_date) as cost_change
FROM product_cost_trends 
WHERE model_name = 'model-A1-001'
ORDER BY extract_date;
```

### 지역별 비교 분석
```sql
-- 지역별 평균 재료비 및 부품별 비중
SELECT 
    region,
    AVG(total_cost) as avg_total,
    AVG(cell_cost / total_cost * 100) as cell_ratio,
    AVG(blu_cost / total_cost * 100) as blu_ratio,
    AVG(circuit_cost / total_cost * 100) as circuit_ratio
FROM product_cost_trends 
GROUP BY region
ORDER BY avg_total DESC;
```

### 이상치 탐지
```sql
-- 전월 대비 재료비 급변동 모델 탐지 (±15% 이상)
WITH monthly_changes AS (
    SELECT 
        model_name,
        extract_date,
        total_cost,
        LAG(total_cost) OVER (PARTITION BY model_name ORDER BY extract_date) as prev_cost,
        ABS((total_cost - LAG(total_cost) OVER (PARTITION BY model_name ORDER BY extract_date)) / LAG(total_cost) OVER (PARTITION BY model_name ORDER BY extract_date) * 100) as change_rate
    FROM product_cost_trends
)
SELECT model_name, extract_date, total_cost, prev_cost, change_rate
FROM monthly_changes 
WHERE change_rate > 15
ORDER BY change_rate DESC;
```

### 부품별 기여도 분석
```sql
-- 시리즈별 부품 비용 구성 분석
SELECT 
    series,
    AVG(cell_cost / total_cost * 100) as cell_contribution,
    AVG(blu_cost / total_cost * 100) as blu_contribution,
    AVG(circuit_cost / total_cost * 100) as circuit_contribution,
    AVG((cell_cost + blu_cost + circuit_cost) / total_cost * 100) as top3_contribution
FROM product_cost_trends 
GROUP BY series
ORDER BY series;
```

---
*생성일: 2024년*  
*문서 버전: 1.0*  
*담당: ABC사 생산기획팀*
*용도: 제품 재료비 트렌드 분석 및 원가 관리*
"""
        
        # 파일 생성 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = f"03_cost_trends_description_{alt_timestamp}.md"
            output_path = os.path.join(refined_folder, alt_filename)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 설명 마크다운 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

def main():
    """메인 실행 함수"""
    generator = CostTrendsGenerator()
    
    try:
        print("=" * 60)
        print("ABC사 제품 재료비 트렌드 분석 데이터 생성 시작")
        print("=" * 60)
        
        # 1. 기존 형식 엑셀 파일 생성 (현재 폴더)
        print("1. 기존 형식 엑셀 파일 생성 중...")
        generated_file = generator.create_excel_file()
        
        # 2. 정제된 엑셀 파일 생성 (상위폴더/refined_excel)
        print("2. 정제된 엑셀 파일 생성 중...")
        refined_file = generator.create_refined_excel_file()
        
        # 3. 설명 마크다운 파일 생성 (상위폴더/refined_excel)
        print("3. 설명 마크다운 파일 생성 중...")
        description_file = generator.create_description_md_file()
        
        print("\n" + "=" * 60)
        print("🎉 ABC사 제품 재료비 트렌드 분석 데이터 생성 완료 🎉")
        print("=" * 60)
        print(f"📊 기존 형식 파일: {generated_file}")
        print(f"📋 정제된 파일: {refined_file}")
        print(f"📝 설명 문서: {description_file}")
        print("\n✅ 생성 내용:")
        print("   • 데이터: 300개 (50개 모델 × 6개월)")
        print("   • 기간: 2025년 2월~7월 (6개월)")
        print("   • 형식: 1줄 헤더 + 재료비 상세 데이터")
        print("   • 부품 구성: 14개 부품별 재료비")
        print("   • 지역: 4개 지역 (북미, 국내, 독일, 영국)")
        print("   • 인치: 15종류 (3~18인치)")
        print("   • 시리즈: 10개 (A1~E2)")
        print("   • 추출 주기: 매월 말일 (월별 동일 날짜)")
        print("   • 설명 문서: SQLite DB 스키마 + 트렌드 분석 쿼리")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
