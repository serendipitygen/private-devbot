"""
ABC사 스마트폰 AI 기능 엑셀 파일 데이터 생성기
3줄로 구성된 제목과 100개 라인의 데이터를 생성합니다.
"""

import random
from datetime import datetime
import os

class AIFeatureDataGenerator:
    def __init__(self):
        # AI 기능 분류 및 기능명 데이터
        self.ai_categories = {
            "카메라 AI": [
                "스마트 HDR", "야간 모드", "인물 사진 모드", "음식 촬영 모드", "자동 장면 인식",
                "실시간 필터", "동영상 안정화", "슬로우 모션", "AI 줌", "배경 흐림"
            ],
            "음성 AI": [
                "음성 인식", "실시간 번역", "음성 명령", "스마트 콜", "소음 제거",
                "음성 메모", "AI 스피커", "음성 잠금해제", "통화 품질 향상", "음성 텍스트 변환"
            ],
            "디스플레이 AI": [
                "적응형 밝기", "블루라이트 필터", "컬러 최적화", "동적 새로고침율", "시선 추적",
                "화면 회전 감지", "스마트 스크린샷", "AI 배경화면", "야간 모드", "에너지 효율"
            ],
            "보안 AI": [
                "얼굴 인식", "지문 인식", "홍채 인식", "음성 인식 보안", "행동 패턴 분석",
                "사기 탐지", "개인정보 보호", "앱 보안", "네트워크 보안", "데이터 암호화"
            ],
            "배터리 AI": [
                "적응형 배터리", "스마트 충전", "전력 최적화", "사용 패턴 학습", "앱별 전력 관리",
                "절전 모드", "충전 예측", "배터리 수명 연장", "열 관리", "무선 충전 최적화"
            ],
            "성능 AI": [
                "CPU 최적화", "메모리 관리", "앱 로딩 가속", "게임 부스터", "멀티태스킹",
                "백그라운드 최적화", "시스템 청소", "스토리지 관리", "네트워크 최적화", "thermal 관리"
            ],
            "개인화 AI": [
                "사용자 습관 학습", "개인 맞춤 UI", "스마트 알림", "컨텐츠 추천", "루틴 자동화",
                "위치 기반 서비스", "시간대별 설정", "앱 추천", "검색 최적화", "개인 비서"
            ],
            "건강 AI": [
                "심박수 모니터링", "수면 분석", "스트레스 측정", "운동량 추적", "칼로리 계산",
                "자세 교정", "눈 건강 관리", "호흡 가이드", "건강 리포트", "의료 데이터 분석"
            ],
            "네트워크 AI": [
                "Wi-Fi 최적화", "5G 연결 관리", "데이터 사용량 예측", "네트워크 속도 향상", "신호 강화",
                "로밍 최적화", "VPN 자동화", "연결 안정성", "대역폭 관리", "트래픽 분석"
            ],
            "엔터테인먼트 AI": [
                "음악 추천", "게임 성능 향상", "동영상 품질 개선", "AR 필터", "VR 최적화",
                "스트리밍 품질", "오디오 향상", "컨텐츠 분류", "미디어 검색", "창작 도구"
            ]
        }
        
        # 도입 시점 옵션
        self.introduction_years = ["21년", "22년", "23년", "24년", "25년"]
        
        # 적용 범위 옵션
        self.application_scope = [
            "24년", "23년 이후 XX로 적용 24년 BB로 적용", "23년", 
            "23년 이후 모든 제품", "24년 플래그십 전용", "23년 이후 프리미엄 라인",
            "24년 하반기", "25년 상반기 예정", "22년 이후 전 모델", "23년 중급형 이상",
            "", "24년 신모델", "유럽/미주"
        ]
        
        # 향지 옵션
        self.regions = [
            "글로벌", "KR/US/UK/AU/FR/GE/IT/ES/PT/BR (10개국)", 
            "KR/US", "글로벌 (중국 제외)", "KR", "US/UK/AU", 
            "유럽/미주", "아시아 태평양", "KR/US/UK/AU/FR/GE/IT/ES/PT/BR (10개국)",
            "", "글로벌", "유럽"
        ]
        
        # 적용 라인업 옵션
        self.lineups = [
            "LINE-P", "LINE-M", "LINEP/LINEM", "LINE-P 전용",
            "LINE-M 전용", "ALL", "플래그십 전용", "프리미엄 이상"
        ]
        
        # 개발 주체 옵션
        self.development_teams = [
            "본사", "해외연 A", "해외연 B", "본사/해외연 A", 
            "해외연 C", "외부 파트너", "본사 주도"
        ]
        
        # 담당부서 옵션
        self.departments = [
            "AI플랫폼팀", "머신러닝팀", "카메라AI팀", "음성인식팀", 
            "컴퓨터비전팀", "자연어처리팀", "딥러닝연구팀", "알고리즘최적화팀",
            "모바일AI팀", "시스템SW팀", "성능최적화팀", "보안기술팀",
            "UX/UI팀", "제품기획팀", "기술전략팀", "품질보증팀"
        ]
        
        # 담당자 옵션
        self.managers = [
            "김태현", "이수민", "박정호", "정예린", "최민준", 
            "장서영", "임도훈", "한지우", "강민석", "윤채영",
            "송준혁", "배소현", "노성민", "서지훈", "오예진",
            "홍준서", "신다은", "조민규", "양서준", "구하은"
        ]
    
    def generate_detailed_description(self, category, feature_name):
        """상세 설명 생성"""
        descriptions = {
            "카메라 AI": [
                f"{feature_name} 기술은 딥러닝 알고리즘을 활용하여 촬영 환경을 실시간으로 분석합니다.",
                "사용자의 촬영 습관을 학습하여 최적의 설정을 자동으로 적용합니다.",
                "다양한 조명 조건에서도 선명하고 자연스러운 이미지를 제공합니다.",
                "전문가 수준의 사진 촬영이 가능하도록 AI가 모든 과정을 지원합니다."
            ],
            "음성 AI": [
                f"{feature_name} 기능은 고도화된 자연어 처리 기술을 기반으로 합니다.",
                "실시간 음성 인식을 통해 사용자의 명령을 정확하게 이해합니다.",
                "노이즈 캔슬링 알고리즘으로 깨끗한 음성 입출력을 보장합니다.",
                "다국어 지원을 통해 글로벌 사용자들에게 편의성을 제공합니다."
            ],
            "디스플레이 AI": [
                f"{feature_name} 시스템은 사용자의 시청 환경을 지능적으로 분석합니다.",
                "주변 광량과 컨텐츠 특성에 맞춰 최적의 화면 설정을 제공합니다.",
                "눈의 피로를 최소화하고 시각적 만족도를 극대화합니다.",
                "배터리 효율성과 화질 품질 간의 완벽한 균형을 유지합니다."
            ],
            "보안 AI": [
                f"{feature_name} 보안 시스템은 생체인식 기술의 정확도를 향상시킵니다.",
                "다중 인증 방식을 결합하여 보안 수준을 한층 강화했습니다.",
                "사용자 행동 패턴 분석을 통해 이상 징후를 조기에 탐지합니다.",
                "개인정보 보호 규정을 완벽하게 준수하며 안전성을 보장합니다."
            ],
            "배터리 AI": [
                f"{feature_name} 관리 시스템은 배터리 수명을 지능적으로 연장합니다.",
                "사용자의 충전 패턴을 학습하여 최적의 충전 전략을 수립합니다.",
                "앱별 전력 소비를 모니터링하여 효율적인 에너지 분배를 실현합니다.",
                "열 관리 알고리즘을 통해 안전하고 지속적인 성능을 유지합니다."
            ],
            "성능 AI": [
                f"{feature_name} 엔진은 시스템 리소스를 효율적으로 관리합니다.",
                "사용 패턴 분석을 통해 필요한 성능을 예측하고 선제적으로 최적화합니다.",
                "멀티태스킹 환경에서도 끊김 없는 사용자 경험을 제공합니다.",
                "하드웨어 특성을 최대한 활용하여 최고의 성능을 구현합니다."
            ],
            "개인화 AI": [
                f"{feature_name} 서비스는 개별 사용자의 선호도를 깊이 있게 학습합니다.",
                "일상 루틴과 사용 패턴을 분석하여 개인 맞춤형 경험을 제공합니다.",
                "프라이버시를 보호하면서도 높은 수준의 개인화를 실현합니다.",
                "사용할수록 더욱 똑똑해지는 적응형 AI 시스템을 구축했습니다."
            ],
            "건강 AI": [
                f"{feature_name} 모니터링은 정확한 생체 신호 측정을 기반으로 합니다.",
                "의료진과의 협력을 통해 검증된 건강 관리 솔루션을 제공합니다.",
                "24시간 연속 모니터링으로 건강 상태 변화를 실시간으로 추적합니다.",
                "개인별 건강 목표 달성을 위한 맞춤형 가이드를 제공합니다."
            ],
            "네트워크 AI": [
                f"{feature_name} 기술은 네트워크 연결 품질을 지속적으로 모니터링합니다.",
                "실시간 트래픽 분석을 통해 최적의 연결 경로를 자동 선택합니다.",
                "다양한 네트워크 환경에서 안정적인 연결성을 보장합니다.",
                "데이터 사용량을 효율적으로 관리하여 비용 절감 효과를 제공합니다."
            ],
            "엔터테인먼트 AI": [
                f"{feature_name} 시스템은 사용자의 취향을 정확하게 분석합니다.",
                "컨텐츠 품질을 실시간으로 향상시켜 최고의 몰입감을 제공합니다.",
                "개인별 선호도에 기반한 지능형 추천 서비스를 구현했습니다.",
                "창작 활동을 지원하는 다양한 AI 도구들을 통합 제공합니다."
            ]
        }
        
        base_descriptions = descriptions.get(category, [
            f"{feature_name} 기능은 최신 AI 기술을 활용합니다.",
            "사용자 경험 향상을 위한 지능형 시스템입니다.",
            "개인화된 서비스 제공을 통해 편의성을 극대화합니다."
        ])
        
        # 2-4개의 설명을 랜덤하게 선택
        selected_descriptions = random.sample(base_descriptions, 
                                           random.randint(2, min(4, len(base_descriptions))))
        return "\n".join(selected_descriptions)
    

    
    def generate_url(self):
        """랜덤 URL 생성"""
        domains = ["abc-tech.com", "ai-research.abc.com", "developer.abc.co.kr", "tech-blog.abc.com"]
        paths = ["features", "ai", "technology", "innovation", "research", "development"]
        pages = ["smartphone-ai", "mobile-intelligence", "ai-features", "smart-functions"]
        
        domain = random.choice(domains)
        path = random.choice(paths)
        page = random.choice(pages)
        
        return f"https://{domain}/{path}/{page}-{random.randint(1000, 9999)}"
    
    def read_existing_file(self, filename="01_3line_titles.xlsx"):
        """기존 엑셀 파일 읽기"""
        import openpyxl
        
        file_path = filename  # 현재 폴더에서 찾기
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except FileNotFoundError:
            return None
    
    def generate_data_row(self):
        """데이터 행 생성 - 기존 엑셀 구조에 맞춤"""
        category = random.choice(list(self.ai_categories.keys()))
        feature_name = random.choice(self.ai_categories[category])
        
        # Resources 데이터 생성
        resources = {
            "NPU": "O" if random.random() > 0.3 else "",
            "CPU": "O" if random.random() > 0.3 else "",
            "DSP": "O" if random.random() > 0.4 else "",
            "Cloud": "O" if random.random() > 0.5 else ""
        }
        
        # 메모리 사용량 (더 현실적인 범위로 조정)
        ddr_usage = random.randint(128, 2048)  # 128KB ~ 2MB
        flash_usage = random.randint(64, 1024)  # 64KB ~ 1MB
        
        row = [
            category,  # A: 분류
            feature_name,  # B: 기능명
            self.generate_detailed_description(category, feature_name),  # C: 상세 설명
            random.choice(self.introduction_years),  # D: 도입 시점
            random.choice(self.application_scope),  # E: 적용 범위 (OSU)
            random.choice(self.regions),  # F: 향지
            resources["NPU"],  # G: NPU
            resources["CPU"],  # H: CPU  
            resources["DSP"],  # I: DSP
            resources["Cloud"],  # J: Cloud
            ddr_usage,  # K: DDR (KB)
            flash_usage,  # L: Flash (KB)
            "O" if random.random() > 0.7 else "",  # M: 생성형
            random.choice(self.lineups),  # N: 적용 라인업
            random.choice(self.development_teams),  # O: 개발 주체
            self.generate_url(),  # P: 관련 자료/링크
            random.choice(self.departments),  # Q: 담당부서
            random.choice(self.managers)  # R: 담당자
        ]
        
        return row
    
    def copy_header_structure(self, source_worksheet, target_worksheet):
        """원본 워크시트의 헤더 구조를 대상 워크시트로 복사"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 1-3행의 헤더 구조 복사
        for row_num in range(1, 4):
            for col_num in range(1, 19):  # A-R 컬럼
                source_cell = source_worksheet.cell(row=row_num, column=col_num)
                target_cell = target_worksheet.cell(row=row_num, column=col_num)
                
                # 셀 값 복사
                target_cell.value = source_cell.value
                
                # 스타일 복사 (폰트, 정렬 등)
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic
                    )
                
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical
                    )
        
        # 병합된 셀 정보 복사
        for merged_range in source_worksheet.merged_cells.ranges:
            if merged_range.max_row <= 3:  # 헤더 영역만
                target_worksheet.merge_cells(str(merged_range))

    def create_excel_file(self, filename="01_3line_titles.xlsx"):
        """기존 엑셀 파일 구조를 유지하면서 새 파일 생성"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # 원본 파일 경로와 새 파일 경로 설정
        original_path = filename  # 현재 폴더의 원본 파일
        new_filename = filename.replace(".xlsx", "_generated.xlsx")
        output_path = new_filename  # 현재 폴더에 새 파일 생성
        
        # 기존 파일 로드 시도
        existing_workbook = self.read_existing_file(filename)
        
        if existing_workbook:
            # 기존 파일이 있는 경우 - 헤더 구조 완전 복사
            original_worksheet = existing_workbook.active
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = original_worksheet.title
            
            # 헤더 구조 복사
            self.copy_header_structure(original_worksheet, worksheet)
            
            start_row = 4
            
        else:
            # 원본 파일이 없는 경우 - 기본 구조 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "AI기능목록"
            
            # 헤더 생성 (이미지 기준으로 정확한 구조)
            # 1행: AI 기능 리스트 (G1에 위치, 병합)
            worksheet.merge_cells('G1:K1')
            worksheet['G1'] = "AI 기능 리스트"
            
            # 2행: 주요 카테고리
            headers_row2 = ["분류", "기능명", "상세 설명", "도입 시점", "적용 범위 (OSU)", "향지", 
                           "Resources", "", "", "", "", "", "생성형", "적용 라인업", "개발 주체", 
                           "관련 자료/링크", "담당부서", "담당자"]
            
            for col, header in enumerate(headers_row2, 1):
                if header:  # 빈 문자열이 아닌 경우만
                    worksheet.cell(row=2, column=col, value=header)
            
            # Resources 병합 (G2:L2)
            worksheet.merge_cells('G2:L2')
            worksheet['G2'] = "Resources"
            
            # 3행: 세부 컬럼
            headers_row3 = ["", "", "", "", "", "", "NPU", "CPU", "DSP", "Cloud", "DDR (KB)", "Flash (KB)", 
                           "", "", "", "", "", ""]
            
            for col, header in enumerate(headers_row3, 1):
                if header:  # 빈 문자열이 아닌 경우만
                    worksheet.cell(row=3, column=col, value=header)
                
            start_row = 4
        
        # 데이터 생성 및 입력
        for i in range(100):
            row_data = self.generate_data_row()
            row_num = start_row + i
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정 (병합된 셀 오류 방지)
        for col_num in range(1, 19):  # A-R 컬럼
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, worksheet.max_row + 1):
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value and not isinstance(cell, MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 10, 최대 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_generated_{alt_timestamp}.xlsx")
            output_path = alt_filename
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_refined_excel_file(self, filename="01_3line_titles.xlsx"):
        """refined 버전의 엑셀 파일 생성 (1줄 헤더)"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        new_filename = filename.replace(".xlsx", "_refined.xlsx")
        output_path = os.path.join(refined_folder, new_filename)
        
        # 새 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "AI기능목록_정제"
        
        # 1줄 헤더 생성
        headers = [
            "AI 기능 분류", "AI 기능명", "AI 기능 상세설명", "AI 기능 도입시점", 
            "AI 기능 적용 범위 (OSU)", "AI 기능 향지", "NPU", "CPU", "DSP", "Cloud", 
            "DDR (KB)", "Flash (KB)", "생성형 여부", "적용 라인업", "개발 주체", 
            "관련 자료/링크", "담당부서", "담당자"
        ]
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 데이터 생성 및 입력
        for i in range(100):
            row_data = self.generate_data_row()
            row_num = i + 2  # 헤더 다음 행부터
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정
        for col_num in range(1, 19):  # A-R 컬럼
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, worksheet.max_row + 1):
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 15, 최대 50)
            adjusted_width = min(max(max_length + 2, 15), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_refined_{alt_timestamp}.xlsx")
            output_path = os.path.join(refined_folder, alt_filename)
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 정제된 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_description_md_file(self):
        """설명 마크다운 파일 생성"""
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        filename = "01_3line_titles_description.md"
        output_path = os.path.join(refined_folder, filename)
        
        md_content = """# ABC사 스마트폰 AI 기능 데이터베이스 명세서

## 1. SQLite 테이블 생성 SQL

```sql
CREATE TABLE ai_functions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ai_category VARCHAR(50) NOT NULL,
    ai_function_name VARCHAR(100) NOT NULL,
    ai_description TEXT NOT NULL,
    introduction_year VARCHAR(10) NOT NULL,
    application_scope VARCHAR(200),
    target_region VARCHAR(200),
    uses_npu CHAR(1) DEFAULT '',
    uses_cpu CHAR(1) DEFAULT '',
    uses_dsp CHAR(1) DEFAULT '',
    uses_cloud CHAR(1) DEFAULT '',
    ddr_usage_kb INTEGER,
    flash_usage_kb INTEGER,
    is_generative CHAR(1) DEFAULT '',
    lineup VARCHAR(100),
    development_team VARCHAR(100),
    related_links VARCHAR(500),
    responsible_department VARCHAR(100),
    responsible_manager VARCHAR(50),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- 인덱스 생성
CREATE INDEX idx_ai_category ON ai_functions(ai_category);
CREATE INDEX idx_introduction_year ON ai_functions(introduction_year);
CREATE INDEX idx_lineup ON ai_functions(lineup);
CREATE INDEX idx_responsible_department ON ai_functions(responsible_department);
```

## 2. 테이블 및 컬럼 상세 설명

### 테이블 개요
- **테이블명**: `ai_functions`
- **목적**: ABC사 스마트폰에 탑재되는 AI 기능들의 상세 정보를 관리
- **데이터 구조**: 각 AI 기능별로 기술적 세부사항, 적용 범위, 담당자 정보 등을 포함

### 컬럼 상세 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `ai_category` | VARCHAR(50) | AI 기능의 대분류 | '카메라 AI', '음성 AI', '보안 AI' |
| `ai_function_name` | VARCHAR(100) | 구체적인 AI 기능명 | '스마트 HDR', '음성 인식', '얼굴 인식' |
| `ai_description` | TEXT | AI 기능의 상세 설명 | '딥러닝 알고리즘을 활용하여...' |
| `introduction_year` | VARCHAR(10) | 해당 기능이 도입된 연도 | '23년', '24년' |
| `application_scope` | VARCHAR(200) | 적용 범위 및 지원 모델 | '24년 플래그십 전용', '23년 이후 모든 제품' |
| `target_region` | VARCHAR(200) | 출시 지역 | '글로벌', 'KR/US/UK' |
| `uses_npu` | CHAR(1) | NPU 사용 여부 | 'O' 또는 공백 |
| `uses_cpu` | CHAR(1) | CPU 사용 여부 | 'O' 또는 공백 |
| `uses_dsp` | CHAR(1) | DSP 사용 여부 | 'O' 또는 공백 |
| `uses_cloud` | CHAR(1) | 클라우드 사용 여부 | 'O' 또는 공백 |
| `ddr_usage_kb` | INTEGER | DDR 메모리 사용량 (KB) | 1024, 2048 |
| `flash_usage_kb` | INTEGER | Flash 메모리 사용량 (KB) | 512, 1024 |
| `is_generative` | CHAR(1) | 생성형 AI 여부 | 'O' 또는 공백 |
| `lineup` | VARCHAR(100) | 적용 제품 라인업 | 'LINE-P', 'LINEP/LINEM' |
| `development_team` | VARCHAR(100) | 개발 주체 | '본사', '해외연 A' |
| `related_links` | VARCHAR(500) | 관련 자료 URL | 'https://abc-tech.com/...' |
| `responsible_department` | VARCHAR(100) | 담당 부서 | 'AI플랫폼팀', '카메라AI팀' |
| `responsible_manager` | VARCHAR(50) | 담당자 | '김태현', '이수민' |
| `created_at` | DATETIME | 데이터 생성일시 | 2024-01-15 10:30:00 |
| `updated_at` | DATETIME | 데이터 수정일시 | 2024-01-15 10:30:00 |

## 3. 예상 질문과 SQL 쿼리

### 자주 묻는 질문들과 해당 SQL 쿼리

| 번호 | 질문 | SQL 쿼리 | 예상 결과 |
|------|------|----------|----------|
| 1 | 카메라 AI 기능은 총 몇 개인가요? | `SELECT COUNT(*) FROM ai_functions WHERE ai_category = '카메라 AI';` | 예: 15개 |
| 2 | 2024년에 도입된 AI 기능들은? | `SELECT ai_function_name, ai_category FROM ai_functions WHERE introduction_year = '24년';` | 스마트 HDR, 실시간 번역 등 |
| 3 | NPU를 사용하는 AI 기능은? | `SELECT ai_function_name FROM ai_functions WHERE uses_npu = 'O';` | 얼굴 인식, 음성 인식 등 |
| 4 | 메모리 사용량이 가장 높은 기능은? | `SELECT ai_function_name, ddr_usage_kb FROM ai_functions ORDER BY ddr_usage_kb DESC LIMIT 5;` | AI 줌: 2048KB 등 |
| 5 | AI플랫폼팀에서 담당하는 기능들은? | `SELECT ai_function_name, ai_category FROM ai_functions WHERE responsible_department = 'AI플랫폼팀';` | 개인화 AI, 성능 최적화 등 |
| 6 | 생성형 AI 기능은 몇 개인가요? | `SELECT COUNT(*) FROM ai_functions WHERE is_generative = 'O';` | 예: 8개 |
| 7 | 글로벌 출시된 기능 중 LINE-P 적용 기능은? | `SELECT ai_function_name FROM ai_functions WHERE target_region = '글로벌' AND lineup LIKE '%LINE-P%';` | 스마트 콜, 야간 모드 등 |
| 8 | 각 카테고리별 평균 메모리 사용량은? | `SELECT ai_category, AVG(ddr_usage_kb) as avg_ddr FROM ai_functions GROUP BY ai_category;` | 카메라 AI: 1024KB 등 |
| 9 | 본사에서 개발한 보안 AI 기능들은? | `SELECT ai_function_name FROM ai_functions WHERE ai_category = '보안 AI' AND development_team = '본사';` | 얼굴 인식, 지문 인식 등 |
| 10 | 2023년 이후 도입된 클라우드 기반 기능은? | `SELECT ai_function_name, introduction_year FROM ai_functions WHERE uses_cloud = 'O' AND introduction_year >= '23년';` | 실시간 번역, 음성 명령 등 |

## 4. 데이터 활용 가이드

### 성능 최적화 팁
- 자주 사용되는 검색 조건에 대해 인덱스 활용
- 대용량 데이터 조회 시 LIMIT 사용 권장
- 메모리 사용량 기반 분석 시 숫자형 컬럼 활용

### 보고서 작성 예시
```sql
-- 연도별 AI 기능 도입 현황
SELECT introduction_year, COUNT(*) as feature_count, 
       AVG(ddr_usage_kb) as avg_memory_usage
FROM ai_functions 
GROUP BY introduction_year 
ORDER BY introduction_year;

-- 부서별 담당 기능 분포
SELECT responsible_department, COUNT(*) as function_count
FROM ai_functions 
GROUP BY responsible_department 
ORDER BY function_count DESC;
```

---
*생성일: 2024년*  
*문서 버전: 1.0*  
*담당: ABC사 AI개발팀*
"""
        
        # 파일 생성 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = f"01_3line_titles_description_{alt_timestamp}.md"
            output_path = os.path.join(refined_folder, alt_filename)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 설명 마크다운 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

def main():
    """메인 실행 함수"""
    generator = AIFeatureDataGenerator()
    
    try:
        print("=" * 60)
        print("ABC사 스마트폰 AI 기능 데이터 생성 시작")
        print("=" * 60)
        
        # 1. 기존 형식 엑셀 파일 생성 (현재 폴더)
        print("1. 기존 형식 엑셀 파일 생성 중...")
        generated_file = generator.create_excel_file()
        
        # 2. 정제된 엑셀 파일 생성 (상위폴더/refined_excel)
        print("2. 정제된 엑셀 파일 생성 중...")
        refined_file = generator.create_refined_excel_file()
        
        # 3. 설명 마크다운 파일 생성 (상위폴더/refined_excel)
        print("3. 설명 마크다운 파일 생성 중...")
        description_file = generator.create_description_md_file()
        
        print("\n" + "=" * 60)
        print("🎉 ABC사 스마트폰 AI 기능 데이터 생성 완료 🎉")
        print("=" * 60)
        print(f"📊 기존 형식 파일: {generated_file}")
        print(f"📋 정제된 파일: {refined_file}")
        print(f"📝 설명 문서: {description_file}")
        print("\n✅ 생성 내용:")
        print("   • 데이터: 100개 AI 기능")
        print("   • 기존 형식: 3줄 헤더 + 병합 셀 유지")
        print("   • 정제 형식: 1줄 헤더 + 깔끔한 구조")
        print("   • 설명 문서: SQLite DB 스키마 + 예제 쿼리")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
