"""
ABC사 스마트폰 AI 기능 제품 적용 현황 엑셀 파일 데이터 생성기
5줄로 구성된 제목과 100개 라인의 데이터를 생성합니다.
AI 기능이 제품 라인업별로 24년/25년 칩셋에 어떻게 적용되었는지 보여주는 테이블입니다.
"""

import random
from datetime import datetime
import os

class AIProductApplicationGenerator:
    def __init__(self):
        # AI 기능 분류 및 기능명 데이터
        self.ai_categories = {
            "카메라 AI": [
                "스마트 HDR", "야간 모드", "인물 사진 모드", "음식 촬영 모드", "자동 장면 인식",
                "실시간 필터", "동영상 안정화", "슬로우 모션", "AI 줌", "배경 흐림"
            ],
            "음성 AI": [
                "음성 인식", "실시간 번역", "음성 명령", "스마트 콜", "소음 제거",
                "음성 메모", "AI 스피커", "음성 잠금해제", "통화 품질 향상", "음성 텍스트 변환"
            ],
            "디스플레이 AI": [
                "적응형 밝기", "블루라이트 필터", "컬러 최적화", "동적 새로고침율", "시선 추적",
                "화면 회전 감지", "스마트 스크린샷", "AI 배경화면", "야간 모드", "에너지 효율"
            ],
            "보안 AI": [
                "얼굴 인식", "지문 인식", "홍채 인식", "음성 인식 보안", "행동 패턴 분석",
                "사기 탐지", "개인정보 보호", "앱 보안", "네트워크 보안", "데이터 암호화"
            ],
            "배터리 AI": [
                "적응형 배터리", "스마트 충전", "전력 최적화", "사용 패턴 학습", "앱별 전력 관리",
                "절전 모드", "충전 예측", "배터리 수명 연장", "열 관리", "무선 충전 최적화"
            ],
            "성능 AI": [
                "CPU 최적화", "메모리 관리", "앱 로딩 가속", "게임 부스터", "멀티태스킹",
                "백그라운드 최적화", "시스템 청소", "스토리지 관리", "네트워크 최적화", "thermal 관리"
            ],
            "개인화 AI": [
                "사용자 습관 학습", "개인 맞춤 UI", "스마트 알림", "컨텐츠 추천", "루틴 자동화",
                "위치 기반 서비스", "시간대별 설정", "앱 추천", "검색 최적화", "개인 비서"
            ],
            "건강 AI": [
                "심박수 모니터링", "수면 분석", "스트레스 측정", "운동량 추적", "칼로리 계산",
                "자세 교정", "눈 건강 관리", "호흡 가이드", "건강 리포트", "의료 데이터 분석"
            ]
        }
        
        # 제품 라인업
        self.lineups = [
            "AB101", "AB102", "AB103", "C20", "D40", 
            "AB104", "AB105", "AB107", "AB108", "XYZ"
        ]
        
        # 칩셋 정보
        self.chipsets = [
            "Chip-A", "Chip-B", "Chip-C", "Chip-D", "Chip-E", 
            "Chip-F", "Chip-G", "Chip-H", "Chip-I", "N/A"
        ]
        
        # 담당부서 옵션
        self.departments = [
            "AI플랫폼팀", "머신러닝팀", "카메라AI팀", "음성인식팀", 
            "컴퓨터비전팀", "자연어처리팀", "딥러닝연구팀", "알고리즘최적화팀",
            "모바일AI팀", "시스템SW팀", "성능최적화팀", "보안기술팀",
            "UX/UI팀", "제품기획팀", "기술전략팀", "품질보증팀"
        ]
        
        # 담당자 옵션
        self.managers = [
            "김태현", "이수민", "박정호", "정예린", "최민준", 
            "장서영", "임도훈", "한지우", "강민석", "윤채영",
            "송준혁", "배소현", "노성민", "서지훈", "오예진",
            "홍준서", "신다은", "조민규", "양서준", "구하은"
        ]
    
    def read_existing_file(self, filename="02_5lines_titles.xlsx"):
        """기존 엑셀 파일 읽기"""
        import openpyxl
        
        file_path = filename  # 현재 폴더에서 찾기
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except FileNotFoundError:
            return None

    def copy_header_structure(self, source_worksheet, target_worksheet):
        """원본 워크시트의 헤더 구조를 대상 워크시트로 복사"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 1-5행의 헤더 구조 복사
        max_col = 2 + len(self.lineups) + 2  # A, B + 라인업들 + 담당부서, 담당자
        for row_num in range(1, 6):
            for col_num in range(1, max_col + 1):
                source_cell = source_worksheet.cell(row=row_num, column=col_num)
                target_cell = target_worksheet.cell(row=row_num, column=col_num)
                
                # 셀 값 복사
                target_cell.value = source_cell.value
                
                # 스타일 복사 (폰트, 정렬 등)
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic
                    )
                
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical
                    )
        
        # 병합된 셀 정보 복사
        for merged_range in source_worksheet.merged_cells.ranges:
            if merged_range.max_row <= 5:  # 헤더 영역만
                target_worksheet.merge_cells(str(merged_range))

    def generate_lineup_chipset_data(self):
        """라인업별 24/25년 칩셋 데이터 생성"""
        lineup_data = {}
        for lineup in self.lineups:
            chipset_25 = random.choice(self.chipsets)
            chipset_24 = random.choice(self.chipsets)
            
            # 일부는 24년에서 25년으로 업그레이드
            if random.random() > 0.3:
                chipset_25 = random.choice([c for c in self.chipsets if c != "N/A"])
            
            lineup_data[lineup] = {
                "25_chipset": chipset_25,
                "24_chipset": chipset_24
            }
        
        return lineup_data

    def generate_data_row(self):
        """데이터 행 생성 - 실제 엑셀 구조에 맞춘 데이터"""
        category = random.choice(list(self.ai_categories.keys()))
        feature_name = random.choice(self.ai_categories[category])
        
        # 기본 행 데이터 (분류, 기능명)
        row = [category, feature_name]
        
        # 각 라인업별로 적용 여부 추가 (O 또는 공백)
        # 실제 이미지를 보면 라인업당 하나의 컬럼씩 있음
        for lineup in self.lineups:
            # 해당 라인업에 적용 여부 (O 또는 공백)
            apply_status = "O" if random.random() > 0.6 else ""
            row.append(apply_status)
        
        # 담당부서와 담당자 추가
        department = random.choice(self.departments)
        manager = random.choice(self.managers)
        sub_manager = random.choice([m for m in self.managers if m != manager])
        
        row.extend([department, f"{manager}({sub_manager})"])
        
        return row

    def create_excel_file(self, filename="02_5lines_titles.xlsx"):
        """기존 엑셀 파일 구조를 유지하면서 새 파일 생성"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # 원본 파일 경로와 새 파일 경로 설정
        original_path = filename  # 현재 폴더의 원본 파일
        new_filename = filename.replace(".xlsx", "_generated.xlsx")
        output_path = new_filename  # 현재 폴더에 새 파일 생성
        
        # 기존 파일 로드 시도
        existing_workbook = self.read_existing_file(filename)
        
        if existing_workbook:
            # 기존 파일이 있는 경우 - 헤더 구조 완전 복사
            original_worksheet = existing_workbook.active
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = original_worksheet.title
            
            # 헤더 구조 복사
            self.copy_header_structure(original_worksheet, worksheet)
            
            start_row = 6  # 헤더가 5줄이므로 6번째 행부터 데이터
            
        else:
            # 원본 파일이 없는 경우 - 기본 구조 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "AI기능적용현황"
            
            # 5줄 헤더 생성
            self.create_default_header(worksheet)
            start_row = 6  # 헤더가 5줄이므로 6번째 행부터 데이터
        
        # 데이터 생성 및 입력
        for i in range(50):  # 라인업이 많아서 50개로 축소
            row_data = self.generate_data_row()
            row_num = start_row + i
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정 (병합된 셀 오류 방지)
        max_col = 2 + len(self.lineups) + 2  # A, B + 라인업들 + 담당부서, 담당자
        for col_num in range(1, max_col + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, worksheet.max_row + 1):
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value and not isinstance(cell, MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 8, 최대 30)
            adjusted_width = min(max(max_length + 2, 8), 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_generated_{alt_timestamp}.xlsx")
            output_path = alt_filename
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_default_header(self, worksheet):
        """기본 5줄 헤더 생성 - 실제 이미지 구조 반영"""
        # 1행: 제목 (전체 병합)
        last_col = 2 + len(self.lineups) + 2  # A, B + 라인업들 + 담당부서, 담당자
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        worksheet['A1'] = "AI 기능 제품 적용 현황 (2025)"
        
        # 2행: 작업 라인업
        worksheet.merge_cells('A2:B2')
        worksheet['A2'] = "작업 라인업"
        
        # 라인업별 헤더 입력
        for i, lineup in enumerate(self.lineups):
            col_idx = 3 + i  # C부터 시작
            worksheet.cell(row=2, column=col_idx, value=lineup)
        
        # 담당부서, 담당자
        dept_col = 3 + len(self.lineups)
        manager_col = dept_col + 1
        worksheet.cell(row=2, column=dept_col, value="담당부서")
        worksheet.cell(row=2, column=manager_col, value="담당자")
        
        # 3행: 분류, 기능명, 라인업, 담당부서, 담당자
        worksheet['A3'] = "분류"
        worksheet['B3'] = "기능명"
        
        for i, lineup in enumerate(self.lineups):
            col_idx = 3 + i
            worksheet.cell(row=3, column=col_idx, value="라인업")
        
        worksheet.cell(row=3, column=dept_col, value="담당부서")
        worksheet.cell(row=3, column=manager_col, value="담당자")
        
        # 4행: '25 SoC, '24 SoC 등
        for i, lineup in enumerate(self.lineups):
            col_idx = 3 + i
            # 25년과 24년 SoC 정보를 랜덤하게 배치
            if i % 2 == 0:
                worksheet.cell(row=4, column=col_idx, value="'25 SoC")
            else:
                worksheet.cell(row=4, column=col_idx, value="'24 SoC")
        
        # 5행: 실제 칩셋명
        for i, lineup in enumerate(self.lineups):
            col_idx = 3 + i
            chipset = random.choice(self.chipsets)
            worksheet.cell(row=5, column=col_idx, value=chipset)

    def create_refined_excel_file(self, filename="02_5lines_titles.xlsx"):
        """refined 버전의 엑셀 파일 생성 (1줄 헤더)"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        new_filename = filename.replace(".xlsx", "_refined.xlsx")
        output_path = os.path.join(refined_folder, new_filename)
        
        # 새 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "AI기능적용현황_정제"
        
        # 1줄 헤더 생성 (실제 구조에 맞게)
        headers = ["AI 기능 분류", "AI 기능명"]
        
        # 각 라인업별로 하나의 컬럼만 추가
        for lineup in self.lineups:
            headers.append(f"{lineup}")
        
        headers.extend(["담당부서", "담당자"])
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 데이터 생성 및 입력
        for i in range(50):
            row_data = self.generate_data_row()
            row_num = i + 2  # 헤더 다음 행부터
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, worksheet.max_row + 1):
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 10, 최대 25)
            adjusted_width = min(max(max_length + 2, 10), 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_refined_{alt_timestamp}.xlsx")
            output_path = os.path.join(refined_folder, alt_filename)
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 정제된 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_description_md_file(self):
        """설명 마크다운 파일 생성"""
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        filename = "02_5line_titles_description.md"
        output_path = os.path.join(refined_folder, filename)
        
        md_content = """# ABC사 스마트폰 AI 기능 제품 적용 현황 데이터베이스 명세서

## 1. SQLite 테이블 생성 SQL

```sql
CREATE TABLE ai_product_application (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ai_category VARCHAR(50) NOT NULL,
    ai_function_name VARCHAR(100) NOT NULL,
    lineup_ab101 CHAR(1) DEFAULT '',
    lineup_ab102 CHAR(1) DEFAULT '',
    lineup_ab103 CHAR(1) DEFAULT '',
    lineup_c20 CHAR(1) DEFAULT '',
    lineup_d40 CHAR(1) DEFAULT '',
    lineup_ab104 CHAR(1) DEFAULT '',
    lineup_ab105 CHAR(1) DEFAULT '',
    lineup_ab107 CHAR(1) DEFAULT '',
    lineup_ab108 CHAR(1) DEFAULT '',
    lineup_xyz CHAR(1) DEFAULT '',
    responsible_department VARCHAR(100),
    responsible_manager VARCHAR(100),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- 칩셋 정보 테이블
CREATE TABLE lineup_chipsets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lineup_name VARCHAR(20) NOT NULL,
    soc_year VARCHAR(10) NOT NULL,
    chipset_name VARCHAR(20),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- 인덱스 생성
CREATE INDEX idx_ai_category_app ON ai_product_application(ai_category);
CREATE INDEX idx_responsible_department_app ON ai_product_application(responsible_department);
CREATE INDEX idx_lineup_chipset ON lineup_chipsets(lineup_name, soc_year);
```

## 2. 테이블 및 컬럼 상세 설명

### 테이블 개요
- **메인 테이블**: `ai_product_application`
- **목적**: ABC사 스마트폰 AI 기능이 각 제품 라인업에 적용된 현황을 연도별로 관리
- **데이터 구조**: 각 AI 기능별로 10개 라인업의 24년/25년 적용 여부를 매트릭스 형태로 관리

### 메인 테이블 컬럼 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `ai_category` | VARCHAR(50) | AI 기능의 대분류 | '카메라 AI', '음성 AI' |
| `ai_function_name` | VARCHAR(100) | 구체적인 AI 기능명 | '스마트 HDR', '음성 인식' |
| `lineup_ab101` | CHAR(1) | AB101 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab102` | CHAR(1) | AB102 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab103` | CHAR(1) | AB103 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_c20` | CHAR(1) | C20 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_d40` | CHAR(1) | D40 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab104` | CHAR(1) | AB104 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab105` | CHAR(1) | AB105 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab107` | CHAR(1) | AB107 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_ab108` | CHAR(1) | AB108 라인업 적용 여부 | 'O' 또는 공백 |
| `lineup_xyz` | CHAR(1) | XYZ 라인업 적용 여부 | 'O' 또는 공백 |
| `responsible_department` | VARCHAR(100) | 담당 부서 | 'AI플랫폼팀', '카메라AI팀' |
| `responsible_manager` | VARCHAR(100) | 담당자 (주담당자, 부담당자) | '김태현(이수민)' |

### 칩셋 정보 테이블 컬럼 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `lineup_name` | VARCHAR(20) | 제품 라인업명 | 'AB101', 'AB102', 'C20' |
| `soc_year` | VARCHAR(10) | SoC 연도 구분 | '25 SoC', '24 SoC' |
| `chipset_name` | VARCHAR(20) | 칩셋명 | 'Chip-A', 'Chip-B', 'N/A' |

## 3. 예상 질문과 SQL 쿼리

### 자주 묻는 질문들과 해당 SQL 쿼리

| 번호 | 질문 | SQL 쿼리 | 예상 결과 |
|------|------|----------|----------|
| 1 | 카메라 AI 기능은 총 몇 개인가요? | `SELECT COUNT(*) FROM ai_product_application WHERE ai_category = '카메라 AI';` | 예: 12개 |
| 2 | AB101 라인업에 적용된 AI 기능들은? | `SELECT ai_function_name, ai_category FROM ai_product_application WHERE lineup_ab101 = 'O';` | 스마트 HDR, 얼굴 인식 등 |
| 3 | 전체 라인업에 공통으로 적용된 AI 기능은? | `SELECT ai_function_name FROM ai_product_application WHERE lineup_ab101='O' AND lineup_ab102='O' AND lineup_c20='O' AND lineup_d40='O';` | 얼굴 인식, 음성 명령 등 |
| 4 | AI플랫폼팀에서 담당하는 기능들은? | `SELECT ai_function_name, ai_category FROM ai_product_application WHERE responsible_department='AI플랫폼팀';` | 개인화 AI, 성능최적화 등 |
| 5 | C20 라인업 전용 AI 기능은? | `SELECT ai_function_name FROM ai_product_application WHERE lineup_c20='O' AND lineup_ab101='' AND lineup_ab102='';` | 프리미엄 카메라 AI 등 |
| 6 | 각 라인업별 AI 기능 탑재 수는? | `SELECT 'AB101' as lineup, COUNT(*) FROM ai_product_application WHERE lineup_ab101='O' UNION ALL SELECT 'AB102', COUNT(*) FROM ai_product_application WHERE lineup_ab102='O';` | AB101: 15개, AB102: 18개 등 |
| 7 | 음성 AI 기능이 적용된 라인업들은? | `SELECT DISTINCT CASE WHEN lineup_ab101='O' THEN 'AB101' END FROM ai_product_application WHERE ai_category='음성 AI' AND lineup_ab101='O';` | AB101, AB104, C20 등 |
| 8 | 각 부서별 담당 기능 수는? | `SELECT responsible_department, COUNT(*) as total_functions FROM ai_product_application GROUP BY responsible_department ORDER BY total_functions DESC;` | AI플랫폼팀: 8개 등 |
| 9 | 특정 칩셋을 사용하는 라인업의 AI 기능은? | `SELECT DISTINCT app.ai_function_name FROM ai_product_application app JOIN lineup_chipsets lc ON lc.chipset_name='Chip-A';` | 고성능 AI 기능들 |
| 10 | 가장 많은 AI 기능이 적용된 라인업은? | `SELECT lineup, feature_count FROM (SELECT COUNT(*) as feature_count, 'AB101' as lineup FROM ai_product_application WHERE lineup_ab101='O') ORDER BY feature_count DESC LIMIT 1;` | AB104: 22개 기능 |

## 4. 데이터 활용 가이드

### 제품 로드맵 분석
```sql
-- 라인업별 AI 기능 적용 현황
SELECT 
    ai_category,
    SUM(CASE WHEN lineup_ab101='O' THEN 1 ELSE 0 END) as ab101_count,
    SUM(CASE WHEN lineup_ab102='O' THEN 1 ELSE 0 END) as ab102_count,
    SUM(CASE WHEN lineup_c20='O' THEN 1 ELSE 0 END) as c20_count
FROM ai_product_application 
GROUP BY ai_category;
```

### 경쟁력 분석
```sql
-- 라인업별 AI 기능 탑재 현황
SELECT 
    'AB101' as lineup,
    COUNT(CASE WHEN lineup_ab101='O' THEN 1 END) as ai_features
FROM ai_product_application
UNION ALL
SELECT 
    'AB102' as lineup,
    COUNT(CASE WHEN lineup_ab102='O' THEN 1 END) as ai_features
FROM ai_product_application
UNION ALL
SELECT 
    'C20' as lineup,
    COUNT(CASE WHEN lineup_c20='O' THEN 1 END) as ai_features
FROM ai_product_application;
```

### 개발 리소스 관리
```sql
-- 부서별 담당 기능과 적용 범위
SELECT 
    responsible_department,
    COUNT(*) as total_functions,
    SUM(CASE WHEN lineup_ab101='O' OR lineup_ab102='O' THEN 1 ELSE 0 END) as premium_lineup_functions,
    SUM(CASE WHEN lineup_c20='O' OR lineup_d40='O' THEN 1 ELSE 0 END) as mid_range_functions
FROM ai_product_application 
GROUP BY responsible_department
ORDER BY total_functions DESC;
```

---
*생성일: 2024년*  
*문서 버전: 1.0*  
*담당: ABC사 AI개발팀*
*용도: 제품 라인업별 AI 기능 적용 현황 관리 및 분석*
"""
        
        # 파일 생성 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = f"02_5line_titles_description_{alt_timestamp}.md"
            output_path = os.path.join(refined_folder, alt_filename)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 설명 마크다운 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

def main():
    """메인 실행 함수"""
    generator = AIProductApplicationGenerator()
    
    try:
        print("=" * 60)
        print("ABC사 스마트폰 AI 기능 제품 적용 현황 데이터 생성 시작")
        print("=" * 60)
        
        # 1. 기존 형식 엑셀 파일 생성 (현재 폴더)
        print("1. 기존 형식 엑셀 파일 생성 중...")
        generated_file = generator.create_excel_file()
        
        # 2. 정제된 엑셀 파일 생성 (상위폴더/refined_excel)
        print("2. 정제된 엑셀 파일 생성 중...")
        refined_file = generator.create_refined_excel_file()
        
        # 3. 설명 마크다운 파일 생성 (상위폴더/refined_excel)
        print("3. 설명 마크다운 파일 생성 중...")
        description_file = generator.create_description_md_file()
        
        print("\n" + "=" * 60)
        print("🎉 ABC사 AI 기능 제품 적용 현황 데이터 생성 완료 🎉")
        print("=" * 60)
        print(f"📊 기존 형식 파일: {generated_file}")
        print(f"📋 정제된 파일: {refined_file}")
        print(f"📝 설명 문서: {description_file}")
        print("\n✅ 생성 내용:")
        print("   • 데이터: 50개 AI 기능")
        print("   • 기존 형식: 5줄 헤더 + 라인업별 매트릭스")
        print("   • 정제 형식: 1줄 헤더 + 라인업별 컬럼")
        print("   • 설명 문서: SQLite DB 스키마 + 제품 로드맵 분석 쿼리")
        print("   • 제품 라인업: 10개 (AB101~XYZ)")
        print("   • 적용 연도: 2024년/2025년 비교")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
